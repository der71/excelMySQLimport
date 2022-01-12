import mysql.connector
import openpyxl
from pathlib import Path
import os

def main():
    all_files = os.listdir("data/")   # imagine you're one directory above test dir

    # connect to MySQL database
    mydb = mysql.connector.connect(
      host="127.0.0.1",
      user="user",
      password="password",
      database="database"
    )

    mycursor = mydb.cursor()

    # Create table if not exists
    sql = """CREATE TABLE IF NOT EXISTS projects (
            file varchar(255),
            Analyte varchar(255),
            HMDB_name varchar(255),
            Chemical varchar(255),
            Monoisotopic FLOAT(24),
            HMDB_id varchar(255),
            KEGG_id varchar(255),
            Group1 varchar(255),
            Group2 varchar(255),
            P_Value FLOAT(24),
            HOLM FLOAT(24),
            BONFERRONI FLOAT(24),
            FDR FLOAT(24),
            VIP_score FLOAT(24),
            Fold_change FLOAT(24),
            Log2_fold_change FLOAT(24)
           );"""
    # print(sql)
    mycursor.execute(sql)
    sql = """DELETE FROM projects WHERE 1"""
    mycursor.execute(sql)

    #Read files
    for file in all_files:
        f_path = 'data/' + file
        print(f_path)
        wb_obj = openpyxl.load_workbook(f_path)

        sheet = wb_obj.active

        # file stats
        print(wb_obj, sheet)
        print(sheet.max_row, sheet.max_column)

        # read header
        col_names = []
        for column in sheet.iter_cols(1, sheet.max_column):
            if column[0].value != "":
                col_names.append(column[0].value)

        print(col_names)
        groups = []
        for col in col_names:
            if col[-3:] == "_01":
                print(col[:-3])
                groups.append(col[:-3])

        Group1, Group2 = groups[0], groups[1]

        n = 0
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                continue
            else:
                n = n + 1
                # put in data in MySQL
                sql = """INSERT INTO projects (file, Analyte, HMDB_name, Chemical,
                                                Monoisotopic, HMDB_id, KEGG_id,
                                                Group1, Group2, P_Value, HOLM,
                                                BONFERRONI, FDR, VIP_score,
                                                Fold_change, Log2_fold_change)
                          VALUES (%s, %s, %s, %s, %s, %s, %s, %s,
                                  %s, %s, %s, %s, %s, %s, %s, %s)"""
                val = (file, row[0], row[1], row[2], row[3], row[4], row[5],
                        Group1, Group2 , row[sheet.max_column - 7],
                        row[sheet.max_column - 6], row[sheet.max_column - 5],
                        row[sheet.max_column - 4], row[sheet.max_column - 3],
                        row[sheet.max_column - 2], row[sheet.max_column - 1]
                      )
                # print(sql, val)
                mycursor.execute(sql, val)
                mydb.commit()
        print(n, "record inserted.")

main()
